import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
import os
import sys

INDIR = sys.argv[1]
DATADIR = sys.argv[2]
# setting = 'MS15'
setting = 'M10'

# # small test cases
# Day = [1,35] 
# BRatio = np.array([0.0,0.002,0.01])
# Setname = ['OZ', 'HOZ']
# Setmeaning = [[0,0], [0,100]]


BRatio = np.array([0.0,0.002,0.01,0.025,0.05,0.075,0.1,0.13])
Day = [1,35,87,95,110,129,176,195,224,268,290,300,330,335,355]

Setname = ['OZ', 'HOZ', 'TOZ', 'TTOZ']
Setmeaning = [[0,0], [0,100], [0,1000], [0,10000]]

# Create list of file names for generation data (set, day, BRatio)
file_names_G = {}
for s in Setname:
    for r in BRatio:
        for d in Day:
            file_names_G[(s,r,d)] = os.path.join(INDIR,'GLPAllbus_24h_'+setting+s+str(d)+'D'+str(r)+'.CSV')
            

# print(file_names_G)

GenData = pd.read_csv(os.path.join(DATADIR,'GenZero.csv'))
# GLP = pd.read_csv(os.path.join(INDIR,'GLPAllbus_24h_MS15OZ1D0.0.CSV'))
# ZeroGLP = GLP[GLP['busID'].isin(GenData['Bus'])]
# ZeroGLPR = ZeroGLP.merge(GenData[['Bus', 'Plb', 'Pub']], left_on='busID', right_on='Bus', how='left')
# ZeroGLPR['GenNor'] = (ZeroGLPR['Gen']-ZeroGLPR['Plb'])/(ZeroGLPR['Pub']-ZeroGLPR['Plb'])
# ZeroGLPR = ZeroGLPR.drop(columns=['Loss', 'LMP', 'Bus'])
# # print(ZeroGLPR.head(30))
# # AVGGenN = ZeroGLPR.groupby('time')['GenNor'].mean().reset_index()
# AVGGenN = ZeroGLPR['GenNor'].mean()
# print(AVGGenN)

def GenNor_AVG(filename, GenDataRange):
    GLP = pd.read_csv(filename)
    ZeroGLP = GLP[GLP['busID'].isin(GenDataRange['Bus'])]
    ZeroGLPR = ZeroGLP.merge(GenDataRange[['Bus', 'Plb', 'Pub']], left_on='busID', right_on='Bus', how='left')
    # ZeroGLPR['GenNor'] = (ZeroGLPR['Gen']-ZeroGLPR['Plb'])/(ZeroGLPR['Pub']-ZeroGLPR['Plb'])
    # # ZeroGLPR = ZeroGLPR.drop(columns=['Loss', 'LMP', 'Bus'])
    # print(ZeroGLPR.head())
    
    # AVGGenN = ZeroGLPR['GenNor'].mean()
    AVGGenN = ZeroGLPR['Gen'].sum()/24*100
    # print(AVGGenN)
    
    return AVGGenN


def create_excel_with_sheets(filename, sheet_names, column_headers, row_headers):
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove the default sheet created by openpyxl
    wb.remove(wb.active)
    
    # Dictionary to store header mappings for quick lookup
    header_mappings = {}

    for sheet_name in sheet_names:
        # Create a new sheet with the given name
        sheet = wb.create_sheet(title=sheet_name)
        
        # Set column headers (in the first row)
        col_mapping = {}
        for col_num, header in enumerate(column_headers, start=2):  # Start from column 2
            sheet.cell(row=1, column=col_num, value=header)
            col_mapping[header] = col_num
        
        # Set row headers (in the first column)
        row_mapping = {}
        for row_num, header in enumerate(row_headers, start=2):  # Start from row 2
            sheet.cell(row=row_num, column=1, value=header)
            row_mapping[header] = row_num
        
        # Store the mappings
        header_mappings[sheet_name] = (row_mapping, col_mapping)
    
    # Save the workbook
    wb.save(filename)

    return header_mappings

# workbook_name = os.path.join(INDIR,'ANorAvgGen.xlsx')
workbook_name = os.path.join(INDIR,'ANorAvgGenZeroMW.xlsx')
header_mappings = create_excel_with_sheets(workbook_name, Setname, Day, BRatio)    
    
# Load the existing workbook
wb = openpyxl.load_workbook(workbook_name)

# Iterate through each (s, r, d) pair, calculate normalized generation (average all generators and 24 hr)
# row index: bidding ratio; column index: day; sheet index: unit generation cost
for (s, r, d), file_name in file_names_G.items():
    if os.path.exists(file_name):        
        # Select the sheet
        sheet = wb[s]
        # Retrieve the row and column mappings for the specific sheet
        row_mapping, col_mapping = header_mappings[s]
        # Get row and column indices from the mappings
        row_index = row_mapping.get(r)
        col_index = col_mapping.get(d)
        
        if row_index is None or col_index is None:
            raise ValueError("Invalid row or column header.")
        
        GenAVG = GenNor_AVG(file_name, GenData)
        sheet.cell(row=row_index, column=col_index, value=GenAVG)    
            
    else:
        print("file doesn't exist", file_name)
# print(df.head())
wb.save(workbook_name)   

   
print("end of ")


